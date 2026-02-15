"""
Build Complete Treasury CMT Dataset
====================================
Combines historical CSV data (1990-2023) with current API data (2024-present)
to create a complete Treasury_CMT_Data_Tool.xlsx starter file.

This creates the INITIAL file with all data through today.
After this, use update_treasury_cmt.py for ongoing updates.

Usage:
    python build_initial_treasury_file.py [--start-date YYYY-MM-DD]
    
Examples:
    python build_initial_treasury_file.py                    # All data from 1990
    python build_initial_treasury_file.py --start-date 2020-01-01  # From 2020 onward
    python build_initial_treasury_file.py --start-date 2022-01-01  # From 2022 onward
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from datetime import datetime
import requests
import xml.etree.ElementTree as ET
import sys
import argparse

def load_csv_data(csv_file, start_date=None):
    """Load historical data from Treasury CSV (1990-2023)"""
    print(f"Loading historical CSV data from {csv_file}...")
    
    df = pd.read_csv(csv_file)
    df['date'] = pd.to_datetime(df['date'], format='%m/%d/%Y')
    
    # Filter by start date if provided
    if start_date:
        original_count = len(df)
        df = df[df['date'] >= start_date]
        filtered_count = len(df)
        print(f"  Filtered from {original_count} to {filtered_count} rows (start: {start_date.date()})")
    
    print(f"  Loaded {len(df)} rows from {df['date'].min().date()} to {df['date'].max().date()}")
    
    return df

def fetch_treasury_api_data(start_year, end_year):
    """Fetch recent data from Treasury API (2024-present)"""
    print(f"\nFetching API data from {start_year} to {end_year}...")
    
    namespace = {
        'd': 'http://schemas.microsoft.com/ado/2007/08/dataservices',
        'm': 'http://schemas.microsoft.com/ado/2007/08/dataservices/metadata',
        'atom': 'http://www.w3.org/2005/Atom'
    }
    
    all_data = []
    
    for year in range(start_year, end_year + 1):
        print(f"  Fetching {year}...", end=' ')
        
        url = "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/pages/xml"
        params = {
            'data': 'daily_treasury_yield_curve',
            'field_tdr_date_value': year
        }
        
        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            
            root = ET.fromstring(response.text)
            entries = root.findall('.//atom:entry', namespace)
            
            for entry in entries:
                content = entry.find('.//m:properties', namespace)
                if content is None:
                    continue
                
                row_data = {}
                
                # Extract date
                date_elem = content.find('.//d:NEW_DATE', namespace)
                if date_elem is not None and date_elem.text:
                    row_data['date'] = pd.to_datetime(date_elem.text[:10])
                
                # Extract all maturities
                maturities = {
                    'BC_1MONTH': '1 mo',
                    'BC_1_5MONTH': '1.5 mo',  # 1.5 month tenor
                    'BC_2MONTH': '2 mo',
                    'BC_3MONTH': '3 mo',
                    'BC_4MONTH': '4 mo',
                    'BC_6MONTH': '6 mo',
                    'BC_1YEAR': '1 yr',
                    'BC_2YEAR': '2 yr',
                    'BC_3YEAR': '3 yr',
                    'BC_5YEAR': '5 yr',
                    'BC_7YEAR': '7 yr',
                    'BC_10YEAR': '10 yr',
                    'BC_20YEAR': '20 yr',
                    'BC_30YEAR': '30 yr'
                }
                
                for xml_name, col_name in maturities.items():
                    elem = content.find(f'.//d:{xml_name}', namespace)
                    if elem is not None and elem.text:
                        try:
                            # API returns percent (5.55), we'll store it as-is
                            # and convert when writing to Excel
                            row_data[col_name] = float(elem.text)
                        except (ValueError, TypeError):
                            row_data[col_name] = None
                
                if 'date' in row_data:
                    all_data.append(row_data)
            
            print(f"✓ ({len(entries)} records)")
            
        except Exception as e:
            print(f"✗ Error: {e}")
    
    if all_data:
        df = pd.DataFrame(all_data)
        print(f"  Total API records: {len(df)}")
        return df
    else:
        return pd.DataFrame()

def create_excel_file(csv_df, api_df, output_file):
    """Create Excel file with combined data"""
    print(f"\nCreating Excel file: {output_file}...")
    
    # Combine dataframes
    print("  Combining CSV and API data...")
    
    # Ensure both have same columns
    all_columns = ['date', '1 mo', '1.5 mo', '2 mo', '3 mo', '4 mo', '6 mo', '1 yr', '2 yr', '3 yr', '5 yr', '7 yr', '10 yr', '20 yr', '30 yr']
    
    for col in all_columns:
        if col not in csv_df.columns:
            csv_df[col] = None
        if len(api_df) > 0 and col not in api_df.columns:
            api_df[col] = None
    
    # Combine
    if len(api_df) > 0:
        combined_df = pd.concat([csv_df, api_df], ignore_index=True)
    else:
        combined_df = csv_df
    
    # Remove duplicates, sort by date (newest first)
    combined_df = combined_df.drop_duplicates(subset=['date'], keep='last')
    combined_df = combined_df.sort_values('date', ascending=False)
    
    print(f"  Total unique records: {len(combined_df)}")
    print(f"  Date range: {combined_df['date'].min().date()} to {combined_df['date'].max().date()}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'CMT Rates'  # Use standard name expected by bootstrap
    
    # Write header
    headers = ['Date', '1Mo', '1.5Mo', '2Mo', '3Mo', '4Mo', '6Mo', '1Yr', '2Yr', '3Yr', '5Yr', '7Yr', '10Yr', '20Yr', '30Yr']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(bold=True)
    
    # Column mapping
    csv_to_excel = {
        '1 mo': 2,    # Column B
        '1.5 mo': 3,  # Column C (1.5Mo)
        '2 mo': 4,    # Column D
        '3 mo': 5,    # Column E
        '4 mo': 6,    # Column F
        '6 mo': 7,    # Column G
        '1 yr': 8,    # Column H
        '2 yr': 9,    # Column I
        '3 yr': 10,   # Column J
        '5 yr': 11,   # Column K
        '7 yr': 12,   # Column L
        '10 yr': 13,  # Column M
        '20 yr': 14,  # Column N
        '30 yr': 15   # Column O
    }
    
    # Write data
    print("  Writing data to Excel...")
    current_row = 2
    
    for idx, row in combined_df.iterrows():
        # Date
        ws.cell(current_row, 1, row['date'])
        ws.cell(current_row, 1).number_format = 'MM/DD/YYYY'
        
        # Rates - convert from percent to decimal
        for csv_col, excel_col in csv_to_excel.items():
            if csv_col in row and pd.notna(row[csv_col]):
                # Convert percent (5.55) to decimal (0.0555)
                decimal_value = float(row[csv_col]) / 100.0
                ws.cell(current_row, excel_col, decimal_value)
                ws.cell(current_row, excel_col).number_format = '0.0000'
        
        current_row += 1
        
        if idx % 1000 == 0:
            print(f"    Processed {idx} rows...")
    
    # Add metadata sheet
    meta_ws = wb.create_sheet('README')
    meta_ws['A1'] = 'Treasury CMT Data Tool'
    meta_ws['A1'].font = Font(bold=True, size=14)
    meta_ws['A3'] = 'Created:'
    meta_ws['B3'] = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
    meta_ws['A4'] = 'Source:'
    meta_ws['B4'] = 'US Treasury Par Yield Curve Rates'
    meta_ws['A5'] = 'Total Records:'
    meta_ws['B5'] = len(combined_df)
    meta_ws['A6'] = 'Date Range:'
    meta_ws['B6'] = f"{combined_df['date'].min().date()} to {combined_df['date'].max().date()}"
    meta_ws['A8'] = 'Note:'
    meta_ws['B8'] = 'Values stored as decimals (0.0555 = 5.55%)'
    meta_ws['A9'] = 'Update:'
    meta_ws['B9'] = 'Use update_treasury_cmt.py to fetch latest data'
    
    # Save
    print("  Saving workbook...")
    wb.save(output_file)
    print(f"✓ Saved to {output_file}")
    
    return True

def main():
    """Main function"""
    
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description='Build complete Treasury CMT dataset from historical CSV + API data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python build_initial_treasury_file.py
    → Load all data from 1990 to present
  
  python build_initial_treasury_file.py --start-date 2020-01-01
    → Load data from 2020-01-01 to present
  
  python build_initial_treasury_file.py --start-date 2022-01-01
    → Load data from 2022-01-01 to present (good for recent analysis)
        """
    )
    parser.add_argument(
        '--start-date',
        type=str,
        default=None,
        help='Start date for data (YYYY-MM-DD format). Default: include all data from 1990'
    )
    
    args = parser.parse_args()
    
    # Parse start date if provided
    start_date = None
    if args.start_date:
        try:
            start_date = pd.to_datetime(args.start_date)
            print(f"Filtering data from {start_date.date()} onward")
        except:
            print(f"Error: Invalid date format '{args.start_date}'. Use YYYY-MM-DD (e.g., 2022-01-01)")
            sys.exit(1)
    
    print("="*70)
    print("Build Complete Treasury CMT Dataset")
    print("="*70)
    print()
    if start_date:
        print(f"Creating Treasury_CMT_Data_Tool.xlsx from {start_date.date()} to today")
    else:
        print("Creating Treasury_CMT_Data_Tool.xlsx with all historical data")
    print()
    
    # Parameters
    csv_file = 'data/par-yield-curve-rates-1990-2023.csv'  # Historical Treasury data
    output_file = 'Treasury_CMT_Data_Tool.xlsx'
    current_year = datetime.now().year
    
    # Step 1: Load CSV (1990-2023)
    try:
        csv_df = load_csv_data(csv_file, start_date=start_date)
    except FileNotFoundError:
        print(f"Error: Could not find {csv_file}")
        print("Please download from: https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_yield_curve")
        sys.exit(1)
    except Exception as e:
        print(f"Error loading CSV: {e}")
        sys.exit(1)
    
    # Step 2: Fetch API data (2024-present)
    # Start from 2024, or last year in CSV + 1
    csv_last_year = csv_df['date'].max().year
    api_start_year = max(2024, csv_last_year + 1)
    
    if api_start_year <= current_year:
        try:
            api_df = fetch_treasury_api_data(api_start_year, current_year)
            # Filter API data by start_date if provided
            if start_date and len(api_df) > 0:
                api_df = api_df[api_df['date'] >= start_date]
        except Exception as e:
            print(f"Warning: Could not fetch API data: {e}")
            print("Proceeding with CSV data only...")
            api_df = pd.DataFrame()
    else:
        print(f"\nNo API data needed (CSV goes to {csv_last_year})")
        api_df = pd.DataFrame()
    
    # Step 3: Create Excel file
    try:
        create_excel_file(csv_df, api_df, output_file)
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        sys.exit(1)
    
    print()
    print("="*70)
    print("✓ SUCCESS!")
    print("="*70)
    print(f"Created: {output_file}")
    print()
    print("Next steps:")
    print("  1. Use this file as your master Treasury data source")
    print("  2. Run update_treasury_cmt.py periodically to add latest data")
    print("  3. Run bootstrap pipeline to generate yield curves")
    print()

if __name__ == '__main__':
    main()
