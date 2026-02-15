"""
Treasury CMT Rate Updater (Python Version)
==========================================
This script updates the Treasury_CMT_Data_Tool.xlsx file with the latest
CMT rates from the U.S. Treasury Department.

Usage:
    python update_treasury_cmt.py

Requirements:
    pip install openpyxl requests
"""

import openpyxl
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
import sys

def fetch_treasury_data(year):
    """Fetch Treasury CMT data for a specific year from the XML API"""
    url = f"https://home.treasury.gov/resource-center/data-chart-center/interest-rates/pages/xml"
    params = {
        'data': 'daily_treasury_yield_curve',
        'field_tdr_date_value': year
    }
    
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data for year {year}: {e}")
        return None

def parse_xml_data(xml_text):
    """Parse XML response and extract CMT rates"""
    namespace = {
        'd': 'http://schemas.microsoft.com/ado/2007/08/dataservices',
        'm': 'http://schemas.microsoft.com/ado/2007/08/dataservices/metadata',
        'atom': 'http://www.w3.org/2005/Atom'
    }
    
    root = ET.fromstring(xml_text)
    entries = root.findall('.//atom:entry', namespace)
    
    data_rows = []
    
    for entry in entries:
        content = entry.find('.//m:properties', namespace)
        if content is None:
            continue
        
        row_data = {}
        
        # Extract date
        date_elem = content.find('.//d:NEW_DATE', namespace)
        if date_elem is not None and date_elem.text:
            row_data['date'] = date_elem.text[:10]  # YYYY-MM-DD format
        
        # Extract all maturities
        maturities = {
            'BC_1MONTH': '1Mo',
            'BC_1_5MONTH': '1.5Mo',  # 1.5 month tenor
            'BC_2MONTH': '2Mo',
            'BC_3MONTH': '3Mo',
            'BC_4MONTH': '4Mo',
            'BC_6MONTH': '6Mo',
            'BC_1YEAR': '1Yr',
            'BC_2YEAR': '2Yr',
            'BC_3YEAR': '3Yr',
            'BC_5YEAR': '5Yr',
            'BC_7YEAR': '7Yr',
            'BC_10YEAR': '10Yr',
            'BC_20YEAR': '20Yr',
            'BC_30YEAR': '30Yr'
        }
        
        # Special handling for 1.5 month - try multiple variations
        variations_1_5mo = ['BC_1.5MONTH', 'BC_1_5MONTH', 'BC_1POINT5MONTH', 
                           'BC_ONEPOINTFIVEMONTH', 'BC_6WEEK']
        
        for xml_name, display_name in maturities.items():
            elem = content.find(f'.//d:{xml_name}', namespace)
            if elem is not None and elem.text:
                try:
                    row_data[display_name] = float(elem.text)
                except (ValueError, TypeError):
                    row_data[display_name] = None
            else:
                row_data[display_name] = None
        
        # Try to find 1.5 month field with various names
        for var_name in variations_1_5mo:
            elem = content.find(f'.//d:{var_name}', namespace)
            if elem is not None and elem.text:
                try:
                    row_data['1.5Mo'] = float(elem.text)
                    break  # Found it, stop trying
                except (ValueError, TypeError):
                    pass
        
        if 'date' in row_data:
            data_rows.append(row_data)
    
    return data_rows

def update_excel_file(filename, start_year, end_year):
    """Update Excel file with Treasury CMT data"""
    
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb['CMT Rates']
        config_ws = wb['Config']
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return False
    
    # Update config
    config_ws['B3'] = start_year
    config_ws['B4'] = end_year
    
    # Clear existing data
    max_row = ws.max_row
    if max_row > 11:
        ws.delete_rows(12, max_row - 11)
    
    # Fetch and write data
    all_data = []
    print(f"Fetching data from {start_year} to {end_year}...")
    
    for year in range(start_year, end_year + 1):
        print(f"  Fetching {year}...", end=' ')
        xml_data = fetch_treasury_data(year)
        
        if xml_data:
            parsed_data = parse_xml_data(xml_data)
            all_data.extend(parsed_data)
            print(f"✓ ({len(parsed_data)} records)")
        else:
            print("✗ Failed")
    
    # Sort by date (newest first)
    all_data.sort(key=lambda x: x['date'], reverse=True)
    
    # Write to Excel
    print(f"\nWriting {len(all_data)} records to Excel...")
    current_row = 12
    
    maturities_order = ['1Mo', '1.5Mo', '2Mo', '3Mo', '4Mo', '6Mo', 
                        '1Yr', '2Yr', '3Yr', '5Yr', '7Yr', '10Yr', '20Yr', '30Yr']
    
    for data in all_data:
        # Write date
        date_obj = datetime.strptime(data['date'], '%Y-%m-%d')
        ws.cell(current_row, 1, date_obj)
        ws.cell(current_row, 1).number_format = 'MM/DD/YYYY'
        
        # Write rates (convert from API percent format to decimal)
        # Treasury API ALWAYS returns values as whole percentages
        # (e.g., 2.22 means 2.22%, not 0.0222)
        # 
        # We convert to decimal (÷100) and store as NUMBER format
        # (NOT percentage format, which would multiply by 100 again!)
        
        for col_idx, maturity in enumerate(maturities_order, start=2):
            value = data.get(maturity)
            if value is not None:
                # Convert from percent (2.22) to decimal (0.0222)
                decimal_value = value / 100.0
                
                ws.cell(current_row, col_idx, decimal_value)
                # Format as decimal NUMBER (e.g., 0.0222), NOT percentage
                ws.cell(current_row, col_idx).number_format = '0.0000'
        
        current_row += 1
    
    # Update last update timestamp
    ws['B6'] = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
    ws['B6'].font = openpyxl.styles.Font(italic=True, color='00008000')  # Green
    
    # Save
    try:
        wb.save(filename)
        print(f"✓ Successfully updated {filename}")
        print(f"  Total records: {len(all_data)}")
        print(f"  Date range: {all_data[-1]['date']} to {all_data[0]['date']}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False

def main():
    """Main function"""
    filename = 'Treasury_CMT_Data_Tool.xlsx'
    
    # Default: Update current year only
    current_year = datetime.now().year
    
    print("=" * 60)
    print("Treasury CMT Rate Updater")
    print("=" * 60)
    print()
    print("Options:")
    print("  1. Quick update (current year only)")
    print("  2. Full historical update (1990 to present)")
    print("  3. Custom year range")
    print()
    
    choice = input("Enter choice (1-3) or press Enter for Quick update: ").strip()
    
    if choice == '2':
        start_year = 1990
        end_year = current_year
    elif choice == '3':
        try:
            start_year = int(input("Enter start year (1990-present): "))
            end_year = int(input("Enter end year (1990-present): "))
        except ValueError:
            print("Invalid input. Using current year.")
            start_year = end_year = current_year
    else:
        start_year = end_year = current_year
    
    print()
    success = update_excel_file(filename, start_year, end_year)
    
    if success:
        print("\n✓ Update completed successfully!")
    else:
        print("\n✗ Update failed!")
        sys.exit(1)

if __name__ == '__main__':
    main()
