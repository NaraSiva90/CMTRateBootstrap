#!/usr/bin/env python3
"""
cmt_bootstrap.py

Bootstrap discount factors / spot (zero) curves from Treasury CMT (coupon-equivalent) rates,
interpreted as par swap rates (modeling assumption).

Schemes:
  - 1: piecewise-constant instantaneous forward per interval (1D Brent on f_i)
  - 2: piecewise-linear instantaneous forward per interval (1D Brent on a_i, b_i via continuity)
  - 3: monotone cubic instantaneous forward per interval (1D Brent on a_i via b(a) reduction)

Short-rate anchor r0:
  Preferred input is a *combined* short-rate history produced by scripts/update_short_rates.py:
      data/short_rates/short_rate_combined.csv
  Columns: Date, Rate, Source (Rate in percent; Source includes SOFR and EFFR).

  If combined is not present, falls back to separate histories:
      data/short_rates/fed_funds_1954_2018.csv   (FRED DFF download, percent)
      data/short_rates/sofr_2018_present.csv     (optional cache)

Missing tenors:
  Skipped (no interpolation).

Outputs:
  - NPZ panel (Option A): stores all dates, tenor-node outputs, parameters, and round-trip validation.
  - Optional Excel output for humans.
"""
from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import brentq

EXP_CAP = 700.0
SOFR_START = pd.Timestamp("2018-04-03")

def safe_exp(x: float) -> float:
    if x > EXP_CAP: return float("inf")
    if x < -EXP_CAP: return 0.0
    return math.exp(x)

def pay_count(t_years: float, nu: int) -> int:
    return int(round(nu * float(t_years) + 1e-9))

def parse_tenor_to_years(label: str) -> float:
    s = re.sub(r"\s+", " ", str(label).strip()).lower()
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(mo|m|month|months)", s)
    if m: return float(m.group(1)) / 12.0
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(yr|y|year|years)", s)
    if m: return float(m.group(1))
    s2 = s.replace(" ", "")
    m = re.fullmatch(r"(\d+(?:\.\d+)?)(mo|m)", s2)
    if m: return float(m.group(1)) / 12.0
    m = re.fullmatch(r"(\d+(?:\.\d+)?)(yr|y)", s2)
    if m: return float(m.group(1))
    raise ValueError(f"Unrecognized tenor label: {label!r}")

def find_header_row(ws, required_first: str = "Date", max_scan_rows: int = 250) -> int:
    for r in range(1, max_scan_rows + 1):
        v = ws.cell(r, 1).value
        if v is None: 
            continue
        if str(v).strip().lower() == required_first.lower():
            return r
    raise ValueError(f"Could not find header row starting with '{required_first}' in sheet {ws.title}")

def read_cmt_rates_from_workbook(path: str, sheet_name: str = "CMT Rates") -> Tuple[List[str], np.ndarray, pd.DataFrame, int, int]:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook missing required sheet {sheet_name!r}")
    ws = wb[sheet_name]

    hr = find_header_row(ws, "Date")
    headers = []
    c = 1
    while True:
        v = ws.cell(hr, c).value
        if v is None or str(v).strip() == "":
            break
        headers.append(str(v).strip())
        c += 1
    tenors = headers[1:]
    T = np.array([parse_tenor_to_years(t) for t in tenors], dtype=float)

    rows = []
    r = hr + 1
    while True:
        d = ws.cell(r, 1).value
        if d is None or str(d).strip() == "":
            break
        vals = [d]
        for j in range(len(tenors)):
            v = ws.cell(r, 2 + j).value
            if v is None or v == "":
                vals.append(np.nan)
            else:
                fv = float(v)
                # Auto-detect format:
                # Treasury rates are typically 0%-25%
                # If value > 1.0, it's in percent form (e.g., 3.72%) → convert to decimal (0.0372)
                # If value <= 1.0, it's already in decimal form (e.g., 0.0372) → keep as-is
                # This handles both Excel formats seamlessly
                vals.append(fv / 100.0 if fv > 1.0 else fv)
        rows.append(vals)
        r += 1

    df = pd.DataFrame(rows, columns=["Date"] + tenors)
    df["Date"] = pd.to_datetime(df["Date"])
    miny = int(df["Date"].dt.year.min())
    maxy = int(df["Date"].dt.year.max())
    return tenors, T, df, miny, maxy

def load_combined_short_rates(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, parse_dates=["Date"])
    if not {"Date","Rate","Source"}.issubset(df.columns):
        raise ValueError(f"Combined short rate file must have Date, Rate, Source: {path}")
    df["Rate_dec"] = pd.to_numeric(df["Rate"], errors="coerce") / 100.0
    df["Source"] = df["Source"].astype(str)
    df = df.dropna(subset=["Date","Rate_dec"]).sort_values("Date")
    return df[["Date","Rate_dec","Source"]]

def load_fed_funds_history(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]
    col_lower = {c.lower(): c for c in df.columns}
    date_col = col_lower.get("date") or col_lower.get("observation_date")
    if date_col is None:
        raise ValueError(f"Fed funds file missing DATE/observation_date: {path}")
    rate_col = col_lower.get("dff") or col_lower.get("value") or col_lower.get("rate")
    if rate_col is None:
        if len(df.columns) == 2:
            rate_col = df.columns[1]
        else:
            raise ValueError(f"Fed funds file missing DFF/value/rate: {path}")
    out = pd.DataFrame({
        "Date": pd.to_datetime(df[date_col]),
        "Rate_dec": pd.to_numeric(df[rate_col], errors="coerce") / 100.0,
        "Source": "EFFR"
    }).dropna(subset=["Date","Rate_dec"]).sort_values("Date")
    return out

def load_sofr_history_optional(path: str) -> pd.DataFrame:
    if not Path(path).exists():
        return pd.DataFrame(columns=["Date","Rate_dec","Source"])
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]
    col_lower = {c.lower(): c for c in df.columns}
    date_col = col_lower.get("date") or col_lower.get("effective date") or list(df.columns)[0]
    rate_col = col_lower.get("sofr") or col_lower.get("rate") or col_lower.get("rate (%)")
    if rate_col is None:
        rate_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    out = pd.DataFrame({
        "Date": pd.to_datetime(df[date_col]),
        "Rate_dec": pd.to_numeric(df[rate_col], errors="coerce") / 100.0,
        "Source": "SOFR"
    }).dropna(subset=["Date","Rate_dec"]).sort_values("Date")
    out = out[out["Date"] >= SOFR_START]
    return out

def build_r0_series(curve_dates: pd.Series, short_df: pd.DataFrame) -> Tuple[np.ndarray, np.ndarray]:
    short_df = short_df.copy()
    short_df["SourceU"] = short_df["Source"].astype(str).str.upper()

    sofr = short_df[short_df["SourceU"] == "SOFR"][["Date","Rate_dec","SourceU"]].set_index("Date").sort_index()
    effr = short_df[short_df["SourceU"] != "SOFR"][["Date","Rate_dec","SourceU"]].set_index("Date").sort_index()

    r0 = np.full(len(curve_dates), np.nan, dtype=float)
    src = np.empty(len(curve_dates), dtype=object)

    for i, d in enumerate(pd.to_datetime(curve_dates)):
        if d >= SOFR_START and len(sofr) > 0:
            s = sofr.loc[:d]
            if len(s) > 0:
                r0[i] = float(s.iloc[-1]["Rate_dec"])
                src[i] = "SOFR"
                continue
        e = effr.loc[:d]
        if len(e) > 0:
            r0[i] = float(e.iloc[-1]["Rate_dec"])
            src[i] = "EFFR"
        else:
            src[i] = "MISSING"
    return r0, src

def annuity_sum(discount_fn: Callable[[float], float], Ti: float, nu: int) -> float:
    m = pay_count(Ti, nu)
    if m <= 0: return 0.0
    s = 0.0
    for k in range(1, m + 1):
        s += discount_fn(k / nu)
    return s / nu

def par_rate(discount_fn: Callable[[float], float], Ti: float, nu: int) -> float:
    A = annuity_sum(discount_fn, Ti, nu)
    if A <= 0 or not math.isfinite(A): return float("nan")
    return (1.0 - discount_fn(Ti)) / A

def int_lin(a: float, b: float, t: float) -> float:
    return 0.5 * a * t * t + b * t

def int_cubic(a: float, b: float, c: float, d: float, t: float) -> float:
    return (a * t**4)/4.0 + (b * t**3)/3.0 + (c * t**2)/2.0 + d * t

@dataclass
class Scheme1Result:
    f: np.ndarray
    P: np.ndarray
    z: np.ndarray
    f_end: np.ndarray
    used_mask: np.ndarray
    warns: List[str]
    discount_fn: Callable[[float], float]

def bootstrap_scheme1(S: np.ndarray, T: np.ndarray, nu: int, f_max: float = 1.0, tol: float = 1e-14) -> Scheme1Result:
    n = len(T)
    used = np.isfinite(S).astype(bool)
    warns: List[str] = []

    f = np.full(n, np.nan); P = np.full(n, np.nan); z = np.full(n, np.nan); f_end = np.full(n, np.nan)

    P_prev, A_prev, T_prev = 1.0, 0.0, 0.0
    seg_starts, seg_ends, seg_f, cum_int = [], [], [], []
    cum_prev = 0.0

    for i in range(n):
        if not used[i]: 
            continue
        Si, Ti = float(S[i]), float(T[i])
        dT = Ti - T_prev
        if dT <= 0: 
            raise ValueError("Non-increasing tenor grid encountered after skipping missing tenors.")

        def shat(fi: float) -> float:
            Pi = P_prev * safe_exp(-fi * dT)
            m_pay = pay_count(dT, nu)
            inc = 0.0
            for k in range(1, m_pay + 1):
                t = k / nu
                inc += safe_exp(-fi * t)
            Ai = A_prev + (P_prev / nu) * inc
            if Ai <= 0 or not math.isfinite(Ai): 
                return float("nan")
            return (1.0 - Pi) / Ai

        def g(fi: float) -> float:
            return shat(fi) - Si

        lo, hi = -f_max, f_max
        gL, gU = g(lo), g(hi)
        if not (math.isfinite(gL) and math.isfinite(gU)) or gL * gU > 0:
            ok = False
            for scale in [2.0, 5.0, 10.0, 20.0]:
                lo2, hi2 = -f_max*scale, f_max*scale
                gL2, gU2 = g(lo2), g(hi2)
                if math.isfinite(gL2) and math.isfinite(gU2) and gL2 * gU2 <= 0:
                    lo, hi = lo2, hi2
                    ok = True
                    break
            if not ok:
                warns.append(f"Scheme1: could not bracket f_i at tenor {i} (T={Ti:g}); leaving NaN.")
                continue

        fi = brentq(g, lo, hi, xtol=tol, rtol=tol, maxiter=1200)

        Pi = P_prev * safe_exp(-fi * dT)
        m_pay = pay_count(dT, nu)
        inc = 0.0
        for k in range(1, m_pay + 1):
            t = k / nu
            inc += safe_exp(-fi * t)
        Ai = A_prev + (P_prev / nu) * inc

        f[i] = fi; P[i] = Pi; z[i] = -math.log(Pi)/Ti; f_end[i] = fi

        seg_starts.append(T_prev); seg_ends.append(Ti); seg_f.append(fi)
        cum_prev += fi * dT
        cum_int.append(cum_prev)

        P_prev, A_prev, T_prev = Pi, Ai, Ti

    seg_starts = np.array(seg_starts, float); seg_ends = np.array(seg_ends, float)
    seg_f = np.array(seg_f, float); cum_int = np.array(cum_int, float)

    def discount_fn(t: float) -> float:
        if len(seg_ends) == 0: 
            return float("nan")
        t = float(t)
        j = int(np.searchsorted(seg_ends, t, side="left"))
        if j == 0:
            integ = seg_f[0] * t
        else:
            integ_prev = float(cum_int[j-1])
            start = float(seg_starts[j])
            integ = integ_prev + seg_f[j] * (t - start)
        return math.exp(-integ)

    return Scheme1Result(f=f, P=P, z=z, f_end=f_end, used_mask=used, warns=warns, discount_fn=discount_fn)

@dataclass
class Scheme2Result:
    a: np.ndarray
    b: np.ndarray
    P: np.ndarray
    z: np.ndarray
    f_end: np.ndarray
    used_mask: np.ndarray
    warns: List[str]
    discount_fn: Callable[[float], float]

def bootstrap_scheme2(S: np.ndarray, T: np.ndarray, r0: float, nu: int, a_max: float = 50.0, tol: float = 1e-14) -> Scheme2Result:
    n = len(T)
    used = np.isfinite(S).astype(bool)
    warns: List[str] = []

    a = np.full(n, np.nan); b = np.full(n, np.nan)
    P = np.full(n, np.nan); z = np.full(n, np.nan); f_end = np.full(n, np.nan)

    P_prev, A_prev, T_prev = 1.0, 0.0, 0.0
    f_prev_end = float(r0)

    seg_starts, seg_ends, seg_a, seg_b, cum_int = [], [], [], [], []
    cum_prev = 0.0

    for i in range(n):
        if not used[i]: 
            continue
        Si, Ti = float(S[i]), float(T[i])
        dT = Ti - T_prev
        if dT <= 0: 
            raise ValueError("Non-increasing tenor grid encountered after skipping missing tenors.")
        bi = f_prev_end

        def shat(ai: float) -> float:
            Pi = P_prev * safe_exp(-int_lin(ai, bi, dT))
            m_pay = pay_count(dT, nu)
            inc = 0.0
            for k in range(1, m_pay + 1):
                t = k / nu
                inc += safe_exp(-int_lin(ai, bi, t))
            Ai = A_prev + (P_prev / nu) * inc
            if Ai <= 0 or not math.isfinite(Ai): 
                return float("nan")
            return (1.0 - Pi) / Ai

        def g(ai: float) -> float:
            return shat(ai) - Si

        lo, hi = -a_max, a_max
        gL, gU = g(lo), g(hi)
        if not (math.isfinite(gL) and math.isfinite(gU)) or gL * gU > 0:
            ok = False
            # CRITICAL FIX: For wide intervals, try NARROWER brackets first, then wider
            # Wide intervals (10+ years) need small slopes, narrow brackets work better
            scales_to_try = [0.001, 0.01, 0.1, 0.5, 1.0, 2.0, 5.0, 10.0, 20.0, 50.0, 100.0]
            for scale in scales_to_try:
                lo2, hi2 = -a_max*scale, a_max*scale
                gL2, gU2 = g(lo2), g(hi2)
                if math.isfinite(gL2) and math.isfinite(gU2) and gL2*gU2 <= 0:
                    lo, hi = lo2, hi2
                    ok = True
                    break
            if not ok:
                warns.append(f"Scheme2: could not bracket a_i at tenor {i} (T={Ti:g}, dT={dT:.3f}); leaving NaN.")
                continue

        ai = brentq(g, lo, hi, xtol=tol, rtol=tol, maxiter=1200)

        Pi = P_prev * safe_exp(-int_lin(ai, bi, dT))
        m_pay = pay_count(dT, nu)
        inc = 0.0
        for k in range(1, m_pay + 1):
            t = k / nu
            inc += safe_exp(-int_lin(ai, bi, t))
        Ai = A_prev + (P_prev / nu) * inc

        a[i], b[i], P[i], z[i] = ai, bi, Pi, -math.log(Pi)/Ti
        f_prev_end = ai * dT + bi
        f_end[i] = f_prev_end

        seg_starts.append(T_prev); seg_ends.append(Ti); seg_a.append(ai); seg_b.append(bi)
        cum_prev += int_lin(ai, bi, dT)
        cum_int.append(cum_prev)

        P_prev, A_prev, T_prev = Pi, Ai, Ti

    seg_starts = np.array(seg_starts, float); seg_ends = np.array(seg_ends, float)
    seg_a = np.array(seg_a, float); seg_b = np.array(seg_b, float); cum_int = np.array(cum_int, float)

    def discount_fn(t: float) -> float:
        if len(seg_ends) == 0: 
            return float("nan")
        t = float(t)
        j = int(np.searchsorted(seg_ends, t, side="left"))
        if j == 0:
            integ = int_lin(seg_a[0], seg_b[0], t)
        else:
            integ_prev = float(cum_int[j-1])
            start = float(seg_starts[j])
            integ = integ_prev + int_lin(seg_a[j], seg_b[j], t - start)
        return math.exp(-integ)

    return Scheme2Result(a=a, b=b, P=P, z=z, f_end=f_end, used_mask=used, warns=warns, discount_fn=discount_fn)

@dataclass
class Scheme3Result:
    a3: np.ndarray
    b3: np.ndarray
    c3: np.ndarray
    d3: np.ndarray
    c_target_next: np.ndarray
    P: np.ndarray
    z: np.ndarray
    f_end: np.ndarray
    used_mask: np.ndarray
    warns: List[str]
    discount_fn: Callable[[float], float]

def bootstrap_scheme3(S: np.ndarray, T: np.ndarray, r0: float, nu: int, a_max: float = 200.0, tol: float = 1e-13) -> Scheme3Result:
    n = len(T)
    used = np.isfinite(S).astype(bool)
    warns: List[str] = []

    s2 = bootstrap_scheme2(S, T, r0=r0, nu=nu)
    idx = np.where(used)[0].tolist()
    if len(idx) == 0:
        nan = np.full(n, np.nan)
        return Scheme3Result(nan, nan, nan, nan, nan, nan, nan, nan, used, ["No valid tenors"], lambda t: float("nan"))

    tau = []
    for k, i in enumerate(idx):
        t_prev = 0.0 if k == 0 else float(T[idx[k-1]])
        tau.append(float(T[i]) - t_prev)
    tau = np.array(tau, float)
    l = np.array([float(s2.a[i]) for i in idx], float)

    m = len(idx)
    c_target = np.zeros(m + 1, float)
    for j in range(m - 1):
        # Handle NaN slopes from Bootstrap 2 failures
        l_j = l[j] if np.isfinite(l[j]) else 0.0
        l_j1 = l[j+1] if np.isfinite(l[j+1]) else 0.0
        
        if l_j * l_j1 > 0:
            c_target[j+1] = (l_j*tau[j] + l_j1*tau[j+1])/(tau[j]+tau[j+1])
        else:
            c_target[j+1] = 0.0
    c_target[m] = 0.0

    a3 = np.full(n, np.nan); b3 = np.full(n, np.nan); c3 = np.full(n, np.nan); d3 = np.full(n, np.nan)
    c_target_next = np.full(n, np.nan)
    P = np.full(n, np.nan); z = np.full(n, np.nan); f_end = np.full(n, np.nan)

    P_prev, A_prev, T_prev = 1.0, 0.0, 0.0
    d_i = float(r0)
    c_i = 0.0

    seg_starts, seg_ends, seg_A, seg_B, seg_C, seg_D, cum_int = [], [], [], [], [], [], []
    cum_prev = 0.0

    for j, i in enumerate(idx):
        Ti, Si = float(T[i]), float(S[i])
        dT = Ti - T_prev
        c_next = float(c_target[j+1])
        c_target_next[i] = c_next

        # CHANGE 1: Scale a_max based on interval width
        # Wide intervals (from missing tenors) need larger cubic coefficients
        interval_scale = max(1.0, dT / 0.15)  # 0.15 yr ≈ normal max spacing
        a_max_scaled = a_max * interval_scale

        def b_of_a(aa: float) -> float:
            return (c_next - c_i - 3.0*aa*dT*dT)/(2.0*dT)

        def residual(aa: float) -> float:
            bb = b_of_a(aa)
            Pi = P_prev * safe_exp(-int_cubic(aa, bb, c_i, d_i, dT))
            m_pay = pay_count(dT, nu)
            inc = 0.0
            for k in range(1, m_pay + 1):
                t = k / nu
                inc += safe_exp(-int_cubic(aa, bb, c_i, d_i, t))
            Ai = A_prev + (P_prev/nu)*inc
            if Ai <= 0 or not math.isfinite(Ai): 
                return float("nan")
            return (1.0 - Pi)/Ai - Si

        lo, hi = -a_max_scaled, a_max_scaled
        gL, gU = residual(lo), residual(hi)
        
        # NUMERICAL SAFETY: If one side is NaN due to overflow, try asymmetric bracketing
        if not math.isfinite(gU) and math.isfinite(gL):
            # Positive side overflows, try smaller positive bound
            for hi_test in [a_max_scaled/10, a_max_scaled/100, 10.0, 1.0, 0.1]:
                gU_test = residual(hi_test)
                if math.isfinite(gU_test):
                    hi, gU = hi_test, gU_test
                    break
        
        if not math.isfinite(gL) and math.isfinite(gU):
            # Negative side overflows, try smaller negative bound  
            for lo_test in [-a_max_scaled/10, -a_max_scaled/100, -10.0, -1.0, -0.1]:
                gL_test = residual(lo_test)
                if math.isfinite(gL_test):
                    lo, gL = lo_test, gL_test
                    break
        
        if not (math.isfinite(gL) and math.isfinite(gU)) or gL*gU > 0:
            ok = False
            # CRITICAL FIX: Try NARROWER brackets first (especially for wide intervals), then wider
            # Wide intervals need small cubic coefficients - narrow brackets work better
            scales_to_try = [0.001, 0.01, 0.1, 0.5, 1.0, 2.0, 5.0, 10.0, 20.0, 50.0, 100.0]
            for scale in scales_to_try:
                lo2, hi2 = -a_max_scaled*scale, a_max_scaled*scale
                gL2, gU2 = residual(lo2), residual(hi2)
                if math.isfinite(gL2) and math.isfinite(gU2) and gL2*gU2 <= 0:
                    lo, hi = lo2, hi2
                    ok = True
                    break
            
            # CHANGE 3: Fallback to piecewise constant forward if cubic fails
            if not ok:
                warns.append(f"Scheme3: cubic failed at tenor {i} (T={Ti:g}, dT={dT:.3f}yr); using constant forward fallback")
                
                # Use constant forward (set a=b=c=0, solve for d)
                # This is the safest fallback - maintains continuity
                aa, bb, c_next_fallback = 0.0, 0.0, 0.0
                
                # Solve for constant forward d_i that matches the par rate
                # Binary search on d_i
                def residual_const(dd: float) -> float:
                    Pi = P_prev * safe_exp(-dd * dT)
                    m_pay = pay_count(dT, nu)
                    inc = 0.0
                    for k in range(1, m_pay + 1):
                        t = k / nu
                        inc += safe_exp(-dd * t)
                    Ai = A_prev + (P_prev/nu)*inc
                    if Ai <= 0 or not math.isfinite(Ai):
                        return float("nan")
                    return (1.0 - Pi)/Ai - Si
                
                # Bracket search for constant forward
                try:
                    dd = brentq(residual_const, -0.10, 0.20, xtol=tol, rtol=tol, maxiter=2000)
                except:
                    # If even constant forward fails, use previous forward
                    dd = d_i
                    warns.append(f"Scheme3: constant forward also failed at tenor {i}; using previous forward")
                
                # Compute discount factor with constant forward
                Pi = P_prev * safe_exp(-dd * dT)
                m_pay = pay_count(dT, nu)
                inc = 0.0
                for k in range(1, m_pay + 1):
                    t = k / nu
                    inc += safe_exp(-dd * t)
                Ai = A_prev + (P_prev/nu)*inc
                
                d_next = dd  # Constant forward extends
                
                a3[i], b3[i], c3[i], d3[i] = aa, bb, c_next_fallback, dd
                P[i], z[i], f_end[i] = Pi, -math.log(Pi)/Ti, d_next
                
                seg_starts.append(T_prev); seg_ends.append(Ti)
                seg_A.append(aa); seg_B.append(bb); seg_C.append(c_next_fallback); seg_D.append(dd)
                cum_prev += dd * dT  # Integral of constant forward
                cum_int.append(cum_prev)
                
                P_prev, A_prev, T_prev = Pi, Ai, Ti
                c_i, d_i = c_next_fallback, d_next
                continue  # Skip Brent solver, move to next tenor

        aa = brentq(residual, lo, hi, xtol=tol, rtol=tol, maxiter=2000)
        bb = b_of_a(aa)

        Pi = P_prev * safe_exp(-int_cubic(aa, bb, c_i, d_i, dT))
        m_pay = pay_count(dT, nu)
        inc = 0.0
        for k in range(1, m_pay + 1):
            t = k / nu
            inc += safe_exp(-int_cubic(aa, bb, c_i, d_i, t))
        Ai = A_prev + (P_prev/nu)*inc

        d_next = aa*dT**3 + bb*dT**2 + c_i*dT + d_i

        a3[i], b3[i], c3[i], d3[i] = aa, bb, c_i, d_i
        P[i], z[i], f_end[i] = Pi, -math.log(Pi)/Ti, d_next

        seg_starts.append(T_prev); seg_ends.append(Ti)
        seg_A.append(aa); seg_B.append(bb); seg_C.append(c_i); seg_D.append(d_i)
        cum_prev += int_cubic(aa, bb, c_i, d_i, dT)
        cum_int.append(cum_prev)

        P_prev, A_prev, T_prev = Pi, Ai, Ti
        d_i, c_i = d_next, c_next

    seg_starts = np.array(seg_starts, float); seg_ends = np.array(seg_ends, float)
    seg_A = np.array(seg_A, float); seg_B = np.array(seg_B, float); seg_C = np.array(seg_C, float); seg_D = np.array(seg_D, float)
    cum_int = np.array(cum_int, float)

    def discount_fn(t: float) -> float:
        if len(seg_ends) == 0: 
            return float("nan")
        t = float(t)
        j = int(np.searchsorted(seg_ends, t, side="left"))
        if j == 0:
            integ = int_cubic(seg_A[0], seg_B[0], seg_C[0], seg_D[0], t)
        else:
            integ_prev = float(cum_int[j-1])
            start = float(seg_starts[j])
            integ = integ_prev + int_cubic(seg_A[j], seg_B[j], seg_C[j], seg_D[j], t - start)
        return math.exp(-integ)

    return Scheme3Result(a3=a3, b3=b3, c3=c3, d3=d3, c_target_next=c_target_next,
                         P=P, z=z, f_end=f_end, used_mask=used, warns=warns, discount_fn=discount_fn)

def save_panel_npz(out_path: str, payload: Dict[str, object]) -> None:
    np_payload = {}
    for k, v in payload.items():
        if isinstance(v, str):
            np_payload[k] = np.array(v, dtype=object)
        elif isinstance(v, list) and (len(v) == 0 or isinstance(v[0], str)):
            np_payload[k] = np.array(v, dtype=object)
        else:
            np_payload[k] = v
    np.savez_compressed(out_path, **np_payload)

def write_excel(out_xlsx: str, tenors: List[str], dates: np.ndarray,
                par_in: np.ndarray, P_T: np.ndarray, z_T: np.ndarray, f_end: np.ndarray,
                par_impl: np.ndarray, err_bp: np.ndarray, maxabs: np.ndarray, rms: np.ndarray,
                method: str, nu: int, r0: np.ndarray, r0_src: np.ndarray,
                params: Dict[str, np.ndarray]) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    def add_df(name: str, df: pd.DataFrame):
        ws = wb.create_sheet(name)
        for j, col in enumerate(df.columns, start=1):
            ws.cell(1, j, col)
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            for j, col in enumerate(df.columns, start=1):
                val = row[col]
                if isinstance(val, float) and np.isnan(val):
                    val = None
                ws.cell(i, j, val)

    df_in = pd.DataFrame(par_in, columns=tenors); df_in.insert(0, "Date", pd.to_datetime(dates))
    add_df("Par Rates (input)", df_in)
    df_P = pd.DataFrame(P_T, columns=tenors); df_P.insert(0, "Date", pd.to_datetime(dates))
    add_df("Discount Factors", df_P)
    df_z = pd.DataFrame(z_T, columns=tenors); df_z.insert(0, "Date", pd.to_datetime(dates))
    add_df("Spot Rates (cc)", df_z)
    df_f = pd.DataFrame(f_end, columns=tenors); df_f.insert(0, "Date", pd.to_datetime(dates))
    add_df("Forward @ TenorEnd", df_f)
    df_impl = pd.DataFrame(par_impl, columns=tenors); df_impl.insert(0, "Date", pd.to_datetime(dates))
    add_df("Par Rates (implied)", df_impl)

    df_err = pd.DataFrame(err_bp, columns=tenors)
    df_err.insert(0, "RMSError_bp", rms); df_err.insert(0, "MaxAbsError_bp", maxabs); df_err.insert(0, "Date", pd.to_datetime(dates))
    add_df("RoundTrip Error (bp)", df_err)

    for pname, parr in params.items():
        df_p = pd.DataFrame(parr, columns=tenors); df_p.insert(0, "Date", pd.to_datetime(dates))
        add_df(f"Param {pname}", df_p)

    meta = pd.DataFrame({"Date": pd.to_datetime(dates), "r0": r0, "r0_source": r0_src.astype(str)})
    add_df("Short Rate Anchor", meta)

    ws = wb.create_sheet("README", 0)
    lines = [
        f"Bootstrapped Treasury CMT curves — {method}",
        "",
        f"Method: {method} (S1=piecewise const fwd; S2=linear fwd; S3=monotone cubic fwd).",
        f"nu (payments/year): {nu}",
        "Short rate anchor from short-rate history (SOFR preferred when available).",
        "Par rates interpreted as swap/par (coupon-equivalent) rates.",
        "Missing tenors: skipped (no interpolation).",
    ]
    for i, ln in enumerate(lines, start=2):
        ws.cell(i, 1, ln)
    ws.column_dimensions["A"].width = 120
    wb.save(out_xlsx)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--scheme", type=int, choices=[1,2,3], required=True)
    ap.add_argument("--nu", type=int, default=24)

    ap.add_argument("--short-rate-combined", default="data/short_rates/short_rate_combined.csv",
                    help="Preferred: output of scripts/update_short_rates.py")
    ap.add_argument("--fed-funds-csv", default="data/short_rates/fed_funds_1954_2018.csv",
                    help="Fallback if combined file missing")
    ap.add_argument("--sofr-csv", default="data/short_rates/sofr_2018_present.csv",
                    help="Fallback if combined file missing")

    ap.add_argument("--out-npz", default=None)
    ap.add_argument("--write-excel", action="store_true")
    ap.add_argument("--out-xlsx", default=None)
    args = ap.parse_args()

    tenors, T, df, miny, maxy = read_cmt_rates_from_workbook(args.workbook)

    combined = Path(args.short_rate_combined)
    if combined.exists():
        short_df = load_combined_short_rates(str(combined))
    else:
        dff = load_fed_funds_history(args.fed_funds_csv)
        sofr = load_sofr_history_optional(args.sofr_csv)
        short_df = pd.concat([dff, sofr], ignore_index=True).sort_values("Date")

    r0, r0_src = build_r0_series(df["Date"], short_df)

    N, K = len(df), len(tenors)
    par_in = df[tenors].to_numpy(float)
    used_mask = np.isfinite(par_in)

    P_T = np.full((N, K), np.nan)
    z_T = np.full((N, K), np.nan)
    f_end = np.full((N, K), np.nan)

    par_impl = np.full((N, K), np.nan)
    err_bp = np.full((N, K), np.nan)
    maxabs = np.full(N, np.nan)
    rms = np.full(N, np.nan)

    params: Dict[str, np.ndarray] = {}
    method = f"S{args.scheme}"
    if args.scheme == 1:
        params["s1_f"] = np.full((N, K), np.nan)
    elif args.scheme == 2:
        params["s2_a"] = np.full((N, K), np.nan)
        params["s2_b"] = np.full((N, K), np.nan)
    else:
        params["s3_a"] = np.full((N, K), np.nan)
        params["s3_b"] = np.full((N, K), np.nan)
        params["s3_c"] = np.full((N, K), np.nan)
        params["s3_d"] = np.full((N, K), np.nan)
        params["s3_c_target_next"] = np.full((N, K), np.nan)

    logs = np.empty(N, dtype=object)
    status = np.zeros(N, dtype=np.int16)

    for i in range(N):
        S = par_in[i, :]
        if not np.isfinite(r0[i]):
            logs[i] = "Missing r0 for date; cannot bootstrap."
            status[i] = 2
            continue

        try:
            if args.scheme == 1:
                res = bootstrap_scheme1(S, T, nu=args.nu)
                params["s1_f"][i, :] = res.f
                P_T[i, :], z_T[i, :], f_end[i, :] = res.P, res.z, res.f_end
                disc_fn, warns = res.discount_fn, res.warns
            elif args.scheme == 2:
                res = bootstrap_scheme2(S, T, r0=float(r0[i]), nu=args.nu)
                params["s2_a"][i, :] = res.a
                params["s2_b"][i, :] = res.b
                P_T[i, :], z_T[i, :], f_end[i, :] = res.P, res.z, res.f_end
                disc_fn, warns = res.discount_fn, res.warns
            else:
                res = bootstrap_scheme3(S, T, r0=float(r0[i]), nu=args.nu)
                params["s3_a"][i, :] = res.a3
                params["s3_b"][i, :] = res.b3
                params["s3_c"][i, :] = res.c3
                params["s3_d"][i, :] = res.d3
                params["s3_c_target_next"][i, :] = res.c_target_next
                P_T[i, :], z_T[i, :], f_end[i, :] = res.P, res.z, res.f_end
                disc_fn, warns = res.discount_fn, res.warns

            for k, Tk in enumerate(T):
                if not np.isfinite(S[k]): 
                    continue
                Si_hat = par_rate(disc_fn, float(Tk), args.nu)
                par_impl[i, k] = Si_hat
                err_bp[i, k] = (Si_hat - float(S[k])) * 1e4

            if np.isfinite(err_bp[i, :]).any():
                maxabs[i] = float(np.nanmax(np.abs(err_bp[i, :])))
                rms[i] = float(math.sqrt(np.nanmean(err_bp[i, :] ** 2)))

            if warns:
                logs[i] = " | ".join(warns)[:30000]
                status[i] = 1
            else:
                logs[i] = ""
                status[i] = 0

        except Exception as e:
            logs[i] = f"Bootstrap failed: {e}"
            status[i] = 2

    dates = df["Date"].to_numpy(dtype="datetime64[D]")

    if args.out_npz is None:
        base = Path(args.workbook).stem
        args.out_npz = str(Path(args.workbook).with_name(f"{base}_curves_{method}_{miny}-{maxy}.npz"))

    payload: Dict[str, object] = {
        "schema_version": "cmt_curve_panel_v1",
        "generator": Path(__file__).name,
        "created_utc": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "method": method,
        "nu": np.array([args.nu], dtype=np.int32),
        "compounding": "cc",
        "dates": dates,
        "tenor_labels": np.array(tenors, dtype=object),
        "tenor_years": np.array(T, dtype=float),
        "par_rates_input": par_in,
        "r0": r0,
        "r0_source": r0_src.astype(object),
        "discount_factors_T": P_T,
        "spot_rates_cc_T": z_T,
        "forward_endpoint_T": f_end,
        "par_rates_implied": par_impl,
        "par_rate_err_bp": err_bp,
        "par_rate_err_maxabs_bp": maxabs,
        "par_rate_err_rms_bp": rms,
        "status_code": status,
        "log_messages": logs.astype(object),
        "tenor_used_mask": used_mask,
    }
    for k, v in params.items():
        payload[k] = v

    save_panel_npz(args.out_npz, payload)
    print(f"Saved NPZ: {args.out_npz}")

    if args.write_excel:
        if args.out_xlsx is None:
            base = Path(args.workbook).stem
            args.out_xlsx = str(Path(args.workbook).with_name(f"{base}_curves_{method}_{miny}-{maxy}.xlsx"))
        write_excel(args.out_xlsx, tenors, dates, par_in, P_T, z_T, f_end, par_impl, err_bp, maxabs, rms,
                    method, args.nu, r0, r0_src, params)
        print(f"Saved Excel: {args.out_xlsx}")

if __name__ == "__main__":
    main()